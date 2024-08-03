import base64
import io
import traceback
import uuid
import zipfile
from flask import Response, abort, current_app
from zipfile import ZipFile
from flask import Flask, jsonify, render_template, request, redirect, url_for, session, send_from_directory
import pandas as pd
import os
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from playsound import playsound
import csv  # Added for CSV file operations
from flask import send_file, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas  # Add this import statement
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from werkzeug.utils import secure_filename
from PIL import Image as PILImage
import pyodbc


from fpdf import FPDF




app = Flask(__name__)

# Constants for file paths

PENDING_SUBMISSIONS_CSV = 'pending_submissions.csv'

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Ensure the upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

SIGNATURE_UPLOAD_FOLDER = "signature_uploads"
app.config["SIGNATURE_UPLOAD_FOLDER"] = SIGNATURE_UPLOAD_FOLDER

# Ensure the upload folder exists
if not os.path.exists(SIGNATURE_UPLOAD_FOLDER):
    os.makedirs(SIGNATURE_UPLOAD_FOLDER)


# Lists to store submissions
submissions_pending_approval = []
approved_submissions = []


# Use the correct path to your 'namewithid.csv' file
NAMWITHID_CSV = 'namewithid.csv'

# Global variable to store the dataframe
df = None

@app.route('/get_names_by_grade/<int:selected_grade>')
def get_names_by_grade(selected_grade):
    # Make sure to load the data if it's not already loaded
    if df is None:
        load_data()
    
    # Filter the dataframe by the selected grade and return the required fields
    filtered_data = df[df['Grade'] == selected_grade][['Learnerid', 'Name']]
    # Convert the filtered data to a list of dictionaries
    names_with_ids = filtered_data.to_dict(orient='records')
    return jsonify(names_with_ids)

def load_data():
    global df
    # Load the data from the CSV file
    df = pd.read_csv(NAMWITHID_CSV)

# Call the load_data function to load the CSV data when the app starts
load_data()

# Function to load pending submissions from a CSV file
def load_pending_submissions():
    pending_submissions = []
    try:
        with open(PENDING_SUBMISSIONS_CSV, mode='r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                pending_submissions.append(row)
    except FileNotFoundError:
        # If the file doesn't exist, start with an empty list
        pass
    return pending_submissions

# Function to save pending submissions to a CSV file
def save_pending_submissions(submissions):
    with open(PENDING_SUBMISSIONS_CSV, mode='w', newline='') as file:
        fieldnames = ["Name", "Grade", "Date", "Notes", 'Username', "Offenses", "LearnerId", "StudentSignature", "TeacherSignature", "WitnessSignature", "offenseId",  # This seems to be repeated for different values, you likely want to use different keys
    "offenseLevel",
    "offenseCode",
    "offenseType",
    "offensePoint"]
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for submission in submissions:
            writer.writerow(submission)

# Load pending submissions when the app starts
submissions_pending_approval = load_pending_submissions()

# Dummy user data for demonstration purposes
users = {
    'JAMESON A -4': {
        'email': 'jamesona@kismet.com',
        'secondary_username': 'Ajameson',
        'password': 'password123',
        'is_admin': True
    },
    'KGONGWANA E -20': {
        'email': 'Kongwana@kismet.com',
        'secondary_username': 'Kongwana',
        'password': 'userpassword',
        'is_admin': False
    },
    'VARGHESE S -16': {
        'email': 'Varghese@kismet.com',
        'secondary_username': 'Varghese',
        'password': 'randompassword1',
        'is_admin': False
    },
    'JOHN B -14': {
        'email': 'John@kismet.com',
        'secondary_username': 'John',
        'password': 'randompassword2',
        'is_admin': False
    },
    'JOUBERT R -2': {
        'email': 'Joubert@kismet.com',
        'secondary_username': 'Joubert',
        'password': 'randompassword3',
        'is_admin': False
    },
    'JAMESON C -12': {
        'email': 'Cjameson@kismet.com',
        'secondary_username': 'Cjameson',
        'password': 'randompassword4',
        'is_admin': False
    },
    'BESTER N -17': {
        'email': 'Bester@kismet.com',
        'secondary_username': 'Bester',
        'password': 'randompassword5',
        'is_admin': False
    },
    'CLAASSENS CH -22': {
        'email': 'Claassens@kismet.com',
        'secondary_username': 'Claassens',
        'password': 'randompassword6',
        'is_admin': False
    },
    'VAN EEDEN B -38': {
        'email': 'Vaneeden@kismet.com',
        'secondary_username': 'Betsie',
        'password': 'randompassword7',
        'is_admin': False
    },
    'GELDENHUYS SM -44': {
        'email': 'Geldenhuys@kismet.com',
        'secondary_username': 'Geldenhuys',
        'password': 'randompassword8',
        'is_admin': False
    },
    'MOSENTHAL V D -46': {
        'email': 'Mosenthal@kismet.com',
        'secondary_username': 'Mosenthal',
        'password': 'randompassword9',
        'is_admin': False
    },
    'BEZUIDENHOUT M -47': {
        'email': 'Martine@kismet.com',
        'secondary_username': 'Martine',
        'password': 'randompassword10',
        'is_admin': False
    },
    'STRYDOM F -48': {
        'email': 'Strydom@kismet.com',
        'secondary_username': 'Strydom',
        'password': 'randompassword11',
        'is_admin': False
    },
    'BRUWER B -51': {
        'email': 'Ekongwana@kismet.com',
        'secondary_username': 'Ekongwana',
        'password': 'randompassword12',
        'is_admin': False
    },
    'FOURIE L -49': {
        'email': 'Ekongwana@kismet.com',
        'secondary_username': 'Ekongwana',
        'password': 'randompassword13',
        'is_admin': False
    },
    'JANKOWITZ J -52': {
        'email': 'Jankowitz@kismet.com',
        'secondary_username': 'Jankowitz',
        'password': 'randompassword14',
        'is_admin': False
    },
    'TAYOB N -53': {
        'email': 'Tayob@kismet.com',
        'secondary_username': 'Nazmeera',
        'password': 'randompassword15',
        'is_admin': False
    },
    'PATTASSERIL BABU B -55': {
        'email': 'Babu@kismet.com',
        'secondary_username': 'Babu',
        'password': 'randompassword16',
        'is_admin': False
    },
    'GELDENHUYS H D -56': {
        'email': 'Geldenhuys@kismet.com',
        'secondary_username': 'Geldenhuys',
        'password': 'randompassword17',
        'is_admin': False
    },
    'GOOSEN C -57': {
        'email': 'Goosen@kismet.com',
        'secondary_username': 'Goosen',
        'password': 'randompassword18',
        'is_admin': False
    },
    'REDIKER L -58': {
        'email': 'Rediker@kismet.com',
        'secondary_username': 'Rediker',
        'password': 'randompassword19',
        'is_admin': False
    },
    'JONKER F -59': {
        'email': 'Jonker@kismet.com',
        'secondary_username': 'Jonker',
        'password': 'randompassword20',
        'is_admin': False
    },
    'DU TOIT WD -60': {
        'email': 'Dutoit@kismet.com',
        'secondary_username': 'Wahnice',
        'password': 'randompassword21',
        'is_admin': False
    },
    'ghoor': {
        'email': 'Ekongwana@kismet.com',
        'secondary_username': 'Ekongwana',
        'password': '786',
        'is_admin': True
    },
    'M Greyling': {
        'email': 'Greyling@kismet.com',
        'secondary_username': 'Greyling',
        'password': '123',
        'is_admin': True
    },
    'MQ': {
        'email': 'Mqcassim@kismet.com',
        'secondary_username': 'Mq',
        'password': '7867',
        'is_admin': True
    }
}

# Set a secret key for session management
app.secret_key = os.urandom(24)

def login_user(identifier, password):
    for username, user_info in users.items():
        if identifier in [user_info.get('email'), user_info.get('secondary_username')]:
            if user_info['password'] == password:
                session['username'] = username  # Store the primary username in the session
                return True
    return False



@app.route('/')
def index():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()
        
        return render_template('index.html', grades=grades, namesByGrade=names_by_grade)
    else:
        # Redirect to the login page if the user is not logged in
        return redirect(url_for('login'))

@app.route('/green')
def green():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()

        # Read the offenses from the CSV file
        greenoffenses_df = pd.read_csv('greencodes.csv')
        
        # Convert the DataFrame to a list of dictionaries for easy handling in the template
        greenoffenses_list = greenoffenses_df.to_dict(orient='records')

        return render_template('green.html', grades=grades, offenses=greenoffenses_list, namesByGrade=names_by_grade)
    else:
        return redirect(url_for('login'))

@app.route('/yellow')
def yellow():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()

        # Read the offenses from the CSV file
        offenses_df = pd.read_csv('yellowcodes.csv')
        
        # Convert the DataFrame to a list of dictionaries for easy handling in the template
        offenses_list = offenses_df.to_dict(orient='records')

        return render_template('yellow.html', grades=grades, offenses=offenses_list, namesByGrade=names_by_grade)
    else:
        return redirect(url_for('login'))
    

@app.route('/pink')
def pink():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()

        # Read the offenses from the CSV file
        pinkoffenses_df = pd.read_csv('pinkcodes.csv')
        
        # Convert the DataFrame to a list of dictionaries for easy handling in the template
        pinkoffenses_list = pinkoffenses_df.to_dict(orient='records')

        return render_template('pink.html', grades=grades, offenses=pinkoffenses_list, namesByGrade=names_by_grade)
    else:
        return redirect(url_for('login'))

@app.route('/super_admin')
def super_admin():
    if 'username' in session:
        username = session['username']
        if users[username]['is_admin']:
            return render_template('super_admin.html')
        else:
            return 'You are not authorized to access this page.'
    return render_template('super_admin.html')

@app.route('/reload_db', methods=['POST'])
def reload_db():
    if 'username' in session:
        username = session['username']
        if users[username]['is_admin']:
            return render_template('super_admin.html')
        else:
            return 'You are not authorized to access this page.'
    # Logic for reloading the database
    return redirect(url_for('super_admin'))

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    
    if request.method == 'POST':
        selected_user = request.form['username']
        new_password = request.form['new_password']
        if selected_user in users:
            users[selected_user]['password'] = new_password
        return redirect(url_for('change_password'))
    return render_template('change_password.html', users=users)

@app.route('/upload_db', methods=['POST'])
def upload_db():
    if 'username' in session:
        username = session['username']
        if users[username]['is_admin']:
            return render_template('super_admin.html')
        else:
            return 'You are not authorized to access this page.'
    # Logic for uploading to the database
    return redirect(url_for('super_admin'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        identifier = request.form['username']  # This can be username, email, or secondary username
        password = request.form['password']

        # Function to validate the user
        def validate_user(identifier, password):
            for username, user_info in users.items():
                if identifier in [username, user_info.get('email'), user_info.get('secondary_username')] and password == user_info['password']:
                    return username  # Return the primary username
            return None

        # Validate the user
        valid_username = validate_user(identifier, password)

        if valid_username:
            session['username'] = valid_username
            return redirect(url_for('index'))

    return render_template('login.html')

@app.route('/select_name', methods=['POST'])
def select_name():
    selected_name = request.form['name']
    
    # Filter data by the selected name and non-empty Date and Offenses
    filtered_data = df[(df['Name'] == selected_name) & ~df['Date'].isnull() & ~df['Offenses'].isnull()]
    
    return render_template('select_name.html', data=filtered_data)


# Define the directory where signature images will be saved
signature_folder = 'signatures'

@app.route('/submit_form', methods=["GET", 'POST'])
def submit_form():
        if request.method == "POST":
            photo = request.files["photo"]

            if photo:
                # Retrieve the selected name from the form data
                selected_name = request.form["name"]

                # Get the current date in the format YYYY-MM-DD
                current_date = datetime.now().strftime("%d-%m-%y")

                # Combine the selected name and current date to create the filename
                raw_filename = f"{selected_name}_{current_date}.jpg"
                filename = secure_filename(raw_filename)

                photo_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)

                # Save the uploaded photo to the specified directory
                photo.save(photo_path)

        if request.method == "POST":
            signature_photo = request.files["signature_photo"]

            if signature_photo:
                # Retrieve the selected name from the form data
                selected_name = request.form["name"]

                # Get the current date in the format YYYY-MM-DD
                current_date = datetime.now().strftime("%d-%m-%y")

                # Combine the selected name and current date to create the filename
                raw_filename = f"{selected_name}_{current_date}.jpg"
                filename = secure_filename(raw_filename)

                signature_photo_path = os.path.join(app.config["SIGNATURE_UPLOAD_FOLDER"], filename)

                # Save the uploaded photo to the specified directory
                signature_photo.save(signature_photo_path)

                
        selected_grade = request.form['grade']
        selected_name = request.form['name']
        selected_offenses = request.form.getlist('offense')
        current_date = datetime.now().strftime('%d-%m-%y')
        notes = request.form['notes']  # Retrieve the notes from the form
        username = session['username']
        selected_learnerid = request.form['learner_id']

        offense_id = request.form.get('Id')
        offense_level = request.form.get('Level')
        offense_code = request.form.get('Code')
        offense_type = request.form.get('Type')
        offense_point = request.form.get('Point')
        
        # Handle the student's signature
        student_signature_data_url = request.form['student_signature']
        if student_signature_data_url:
            # Generate a unique filename for the student's signature image
            student_signature_filename = f'{selected_name}_{current_date}.png'
            student_signature_path = os.path.join(signature_folder, student_signature_filename)

            # Convert the data URL back to an image and save it as a PNG
            student_signature_image_data = student_signature_data_url.split(',')[1]
            student_signature_image_binary = base64.b64decode(student_signature_image_data)
            with open(student_signature_path, 'wb') as student_signature_file:
                student_signature_file.write(student_signature_image_binary)

        # # Generate a unique filename for the digital signature image
        # student_signature_filename = f'{selected_name}.png'
        # signature_path1 = os.path.join(signature_folder, student_signature_filename)

        # # Create an image with the selected name as signature
        # img1 = Image.new('RGB', (650, 250), color = (255, 255, 255))
        # d = ImageDraw.Draw(img1)
        # font = ImageFont.truetype("static/Andina Demo.otf", 48)  # Specify the path to a signature-style font file
        # d.text((10,10), selected_name, font=font, fill=(0,0,0))

        # # Generate a unique filename for the teacher digital signature image
        # teacher_signature_filename = f'{username}.png'
        # signature_path = os.path.join(signature_folder, teacher_signature_filename)

        # # Create an image with the selected name as signature
        # img = Image.new('RGB', (650, 250), color = (255, 255, 255))
        # d1 = ImageDraw.Draw(img)
        # font = ImageFont.truetype("static/Andina Demo.otf", 48)  # Specify the path to a signature-style font file
        # d1.text((10,10), username, font=font, fill=(0,0,0))

        # # Save the signature image

        # img1.save(signature_path1)
        # img.save(signature_path)

        # Handle the teacher's signature (similar code as above)
        teacher_signature_data_url = request.form['teacher_signature']
        if teacher_signature_data_url:
            teacher_signature_filename = f'{username}_{current_date}.png'
            teacher_signature_path = os.path.join(signature_folder, teacher_signature_filename)
            teacher_signature_image_data = teacher_signature_data_url.split(',')[1]
            teacher_signature_image_binary = base64.b64decode(teacher_signature_image_data)
            with open(teacher_signature_path, 'wb') as teacher_signature_file:
                teacher_signature_file.write(teacher_signature_image_binary)


        # Handle the witness's signature (similar code as above)
        witness_signature_data_url = request.form['witness_signature']
        if witness_signature_data_url:
            witness_signature_filename = f'w_{selected_name}_{current_date}.png'
            witness_signature_path = os.path.join(signature_folder, witness_signature_filename)
            witness_signature_image_data = witness_signature_data_url.split(',')[1]
            witness_signature_image_binary = base64.b64decode(witness_signature_image_data)
            with open(witness_signature_path, 'wb') as witness_signature_file:
                witness_signature_file.write(witness_signature_image_binary)


        # Create a new submission dictionary
        new_submission = {
            'Name': selected_name,
            'Grade': selected_grade,
            'Date': current_date,
            'Offenses': ', '.join(selected_offenses),
            'Notes': notes,
            'Username' : username,
            'LearnerId' : selected_learnerid,
            'offenseId' : offense_id,
            "offenseLevel": offense_level,
            "offenseCode": offense_code,
            "offenseType": offense_type,
            "offensePoint": offense_point


        }

        # Append the submission to the pending approval list
        submissions_pending_approval.append(new_submission)

        # Save pending submissions to the CSV file
        save_pending_submissions(submissions_pending_approval)

        # Notify with a sound
        playsound('static/ding.mp3')

        return redirect(url_for('index'))
    
    


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/admin')
def admin():
    if 'username' in session:
        username = session['username']
        if users[username]['is_admin']:
            return render_template('admin.html', submissions=submissions_pending_approval)
        else:
            return 'You are not authorized to access this page.'
    return redirect(url_for('login'))

@app.route('/approve_submission/<int:index>')
def approve_submission(index):
    if 0 <= index < len(submissions_pending_approval):
        submission = submissions_pending_approval.pop(index)
        approved_submissions.append(submission)

        # Create directory structure for the grade
        grade_folder = os.path.join('Demerits', f'Grade_{submission["Grade"]}', f'{submission["Name"]}')
        os.makedirs(grade_folder, exist_ok=True)

        # Get the current time in HH:MM format
        current_time = datetime.now().strftime("%H:%M")

        logo_image = Image.open('static/Logo.png')
        logo_image = logo_image.convert('RGB')  # Convert to RGB format (optional)
        logo_image.save('static/Logo_non_interlaced1.png', 'PNG', optimize=True)

        logo_filename = 'static/Logo_non_interlaced1.png'
        logo_width = 130  # Width of the logo in millimeters
        logo_height = 40  # Height of the logo in millimeters


        # Use the current time in the filename
        pdf_filename = f'{submission["Name"].replace(" ", "_")}_{submission["Date"]}_{current_time}.pdf'

        pdf_path = os.path.join(grade_folder, pdf_filename)

        # Create a PDF with A4 size (210x297 mm) in landscape mode
        pdf = FPDF(format='A4', unit='mm')
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)  # Automatic page break with a 15mm margin
        pdf.set_font("Arial", size=12)
        
        # Calculate the x-coordinate to center the logo horizontally
        logo_x = (pdf.w - logo_width) / 2

        # Load and embed the logo image
        pdf.image(logo_filename, x=logo_x, y=pdf.get_y(), w=logo_width, h=logo_height)  # Adjust position and size

        # Move down to create space for the logo
        pdf.ln(logo_height)


        # Adjusted positions and sizes for A4 portrait
        pdf.cell(0, 10, txt="Demerits Form", ln=True, align='C')  # Background fill
        pdf.ln(10)  # Add some space

        # Left Column
        pdf.set_fill_color(200, 200, 200)  # Background fill color
        column_width = 145  # Adjust the column width as needed
        column_width1 = 25  # Adjust the column width as needed

        pdf.cell(column_width1, 10, txt="Name:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Name'], ln=True)

        pdf.cell(column_width1, 10, txt="Grade:", fill=True)
        pdf.cell(column_width, 10, txt=str(submission['Grade']), ln=True)

        pdf.cell(column_width1, 10, txt="Date:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Date'], ln=True)

        pdf.cell(column_width1, 10, txt="Notes:", fill=True)
        # Use multi_cell to handle lengthy notes
        pdf.multi_cell(column_width, 10, txt=submission['Notes'])  # Automatically adjusts height


        pdf.cell(column_width1, 10, txt="Educator:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Username'], ln=True)

        

        #pdf.cell(column_width1, 10, txt="Offenses:", fill=True)
        #pdf.cell(column_width, 10, txt=submission['Offenses'], ln=True)  # Multiline cell for offenses

        offense_level = submission["offenseLevel"]
    
        # Print the offense level to debug
        print(f"Offense level: {offense_level}")

        if offense_level == '1':
            color = (255, 255, 0)  # Yellow color for level 1
            print("Setting color to Yellow for level 1")
        elif offense_level == '2':
            color = (255, 182, 193)  # Pink color for level 2
            print("Setting color to Pink for level 2")
        elif offense_level == '3':
            color = (128, 0, 128)  # Purple color for level 3
            print("Setting color to Purple for level 3")
        else:
            color = (255, 255, 255)  # Default white color for other levels
            print("Offense level not recognized, defaulting to white")

        # Apply the fill color
        pdf.set_fill_color(*color)

        # Create cells
        pdf.cell(column_width1, 10, txt="Offenses:", fill=True)
        pdf.multi_cell(column_width, 10, txt=submission['Offenses'], fill=True)  # Automatically adjusts height


        # Reset the fill color to the default (gray) for other cells
        pdf.set_fill_color(200, 200, 200)
        print("Reset fill color to default gray")

        pdf.cell(column_width1, 10, txt="Points:", fill=True)
        pdf.cell(column_width, 10, txt=submission['offensePoint'], ln=True)

        pdf.ln(10)  # Add some space

        # # Title for the student's signature cell
        # pdf.set_font("Arial", style='B', size=12)
        # pdf.cell(0, 10, txt="Student Signature", ln=True, align='C')

        # # Create a cell for the student's signature
        # pdf.set_fill_color(255, 255, 255)  # White background
        # pdf.cell(0, 30, txt="", border=1, ln=True)  # Cell for the signature with borders

        # # Load and embed the student's signature image within the cell
        # student_signature_filename = f'{submission["Name"]}_{submission["Date"]}.png'
        # student_signature_path = os.path.join(signature_folder, student_signature_filename)
        # pdf.image(student_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() - 25, w=0, h=20)  # Adjust position and size

        # # Title for the teacher's signature cell
        # pdf.set_font("Arial", style='B', size=12)
        # pdf.cell(0, 10, txt="Teacher Signature", ln=True, align='C')

        # # Create a cell for the teacher's signature
        # pdf.set_fill_color(255, 255, 255)  # White background
        # pdf.cell(0, 30, txt="", border=1, ln=True)  # Cell for the signature with borders

        # # Load and embed the teacher's signature image within the cell
        # teacher_signature_filename = f'{submission["Username"]}_{submission["Date"]}.png'
        # teacher_signature_path = os.path.join(signature_folder, teacher_signature_filename)
        # pdf.image(teacher_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() - 25, w=0, h=20)  # Adjust position and size

        # Set up for signatures in columns
        signature_height = 30  # Height for signature cells
        signature_width = 60   # Width for each signature column

        # Title for the signatures section
        pdf.set_font("Arial", style='B', size=12)
        pdf.cell(0, 10, txt="Signatures", ln=True, align='C')
        pdf.ln(5)  # Add some space before the signatures

        # Create headers for the signatures
        pdf.set_fill_color(255, 255, 255)  # White background for signature cells

        # Headers for signature columns
        pdf.cell(signature_width, 10, txt="Student", border=1, ln=0, align='C', fill=True)
        pdf.cell(signature_width, 10, txt="Teacher", border=1, ln=0, align='C', fill=True)
        pdf.cell(signature_width, 10, txt="Witness", border=1, ln=1, align='C', fill=True)

        # Create cells for the signatures
        pdf.cell(signature_width, signature_height, txt="", border=1, ln=0, align='C', fill=True)
        pdf.cell(signature_width, signature_height, txt="", border=1, ln=0, align='C', fill=True)
        pdf.cell(signature_width, signature_height, txt="", border=1, ln=1, align='C', fill=True)

        # Move the cursor back to the start of the signatures row
        pdf.set_y(pdf.get_y() - signature_height)

        # Load and embed the student signature image within the cell
        student_signature_filename = f'{submission["Name"]}_{submission["Date"]}.png'
        student_signature_path = os.path.join(signature_folder, student_signature_filename)
        if os.path.exists(student_signature_path):
            pdf.image(student_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() + 5, w=signature_width - 10, h=20)

        # Move the cursor to the next cell
        pdf.set_x(pdf.get_x() + signature_width)

        # Load and embed the teacher signature image within the cell
        teacher_signature_filename = f'{submission["Username"]}_{submission["Date"]}.png'
        teacher_signature_path = os.path.join(signature_folder, teacher_signature_filename)
        if os.path.exists(teacher_signature_path):
            pdf.image(teacher_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() + 5, w=signature_width - 10, h=20)

        # Move the cursor to the next cell
        pdf.set_x(pdf.get_x() + signature_width)

        # Optional Witness Signature
        witness_signature_filename = f'w_{submission["Name"]}_{submission["Date"]}.png'
        witness_signature_path = os.path.join(signature_folder, witness_signature_filename)
        if os.path.exists(witness_signature_path):
            pdf.image(witness_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() + 5, w=signature_width - 10, h=20)

        pdf.ln(5)  # Add some space after the signatures


        # # Title for the witness's signature
        # pdf.set_font("Arial", style='B', size=12)
        # pdf.cell(20190, 10, txt="Witness Signature", ln=True, align='C')

        # # Create a box for the witness's signature
        # pdf.set_fill_color(255, 255, 255)  # White background
        # pdf.rect(10, pdf.get_y(), 190, 40, style='F')  # Rectangle for the signature

        # # Load and embed the witness's signature image in the box
        # witness_signature_filename = f'witness_signature_{submission["Date"]}.png'
        # witness_signature_path = os.path.join(signature_folder, witness_signature_filename)
        # pdf.image(witness_signature_path, x=15, y=pdf.get_y() + 5, w=180, h=30)  # Adjust position and size


        # Extract the image filename from the submission data
        image_filename = f'{submission["Name"].replace(" ", "_")}_{submission["Date"]}.jpg'
        image_path = os.path.join(app.config["UPLOAD_FOLDER"], image_filename)

        


        # Check if the image exists
        if os.path.exists(image_path):
            # Add a new page to the PDF
            pdf.add_page()

            # Use PIL to get image dimensions
            with PILImage.open(image_path) as img:
                img_width, img_height = img.size

            # A4 dimensions in mm
            a4_width_mm = 210
            a4_height_mm = 297

            # Calculate scaling factors to fit half of the page
            width_scale = (a4_width_mm / 2) / img_width
            height_scale = (a4_height_mm / 2) / img_height

            # Use the minimum scale factor to maintain the aspect ratio
            scale_factor = min(width_scale, height_scale)

            # Calculate the new image dimensions
            new_img_width = img_width * scale_factor
            new_img_height = img_height * scale_factor

            # Center the image on the half-page (top half)
            x_offset = (a4_width_mm - new_img_width) / 2
            y_offset = (a4_height_mm / 4) - (new_img_height / 2)

            # Embed the image in the new page
            pdf.image(image_path, x=x_offset, y=y_offset, w=new_img_width, h=new_img_height)

        # Save the PDF to the file
        pdf.output(pdf_path)

        # Parse the original date assuming it's in the format "dd/mm/yy"
        original_date = datetime.strptime(submission['Date'], '%d-%m-%y')

        # Format the date into the required formats
        formatted_date = original_date.strftime('%d/%m/%y') + ' 00:00:00'  # Adding the '00:00:00' part manually
        formatted_month = original_date.strftime('%m')
        formatted_year = original_date.strftime('%Y')

        # Now, map the submission data to your new CSV headers
        mapped_submission = {
            'id': '',  # Assuming you need to generate or have an ID
            'Learnerid': submission['LearnerId'],  # Replace with actual key if exists
            'Date': formatted_date,
            'Comment': submission['Notes'],
            'LevelMisconduct': submission['offenseLevel'],
            'MisconductCode': submission['offenseCode'],
            'MisconductDescription': submission['Offenses'],
            'ActionLevel' : '',
            'ActionCode' : '',
            'ActionDescription': '',
            'DisciplinedBy': '',
            'AuthorisedBy': submission['Username'],
            'Agency': '',
            'Suspension': '0',
            'Option': '',
            'ExpulsionDate': '',
            'Month': formatted_month,
            'RecommendedExpulsion': '',
            'Datayear': formatted_year,
            'Demerit' : submission['offensePoint'],
            'Merit': '0',
            'Type': 'Demerit'
        }

        # Define the path to the CSV file
        csv_file_path = 'mdb-import.csv'

        # Check if the file exists and if not, write the header
        file_exists = os.path.isfile(csv_file_path)

        with open(csv_file_path, 'a', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=mapped_submission.keys())
            
            # Write the header only if the file does not exist
            if not file_exists:
                writer.writeheader()
            
            writer.writerow(mapped_submission)

        # Save pending submissions to the CSV file
        save_pending_submissions(submissions_pending_approval)

    return redirect(url_for('admin'))

@app.route('/reject_submission/<int:index>')
def reject_submission(index):
    if 0 <= index < len(submissions_pending_approval):
        submission = submissions_pending_approval.pop(index)
        # Perform any rejection action here (e.g., delete the submission)
        # Save pending submissions to the CSV file
        save_pending_submissions(submissions_pending_approval)
        # Notify with a sound

    return redirect(url_for('admin'))

@app.route('/reset_mdb', methods=['POST'])
def reset_mdb():
    # Define the path to the CSV file
    csv_file_path = 'mdb-import.csv'
    
    # Open the file in write mode to clear its contents
    with open(csv_file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        # Write the headers to the CSV file
        writer.writerow([
            'id', 'Learnerid', 'Date', 'Comment', 'LevelMisconduct', 'MisconductCode',
            'MisconductDescription', 'ActionLevel', 'ActionCode', 'ActionDescription',
            'DisciplinedBy', 'AuthorisedBy', 'Agency', 'Suspension', 'Option',
            'ExpulsionDate', 'Month', 'RecommendedExpulsion', 'Datayear',
            'Demerit', 'Merit', 'Type'
        ])
    
    # Redirect to the admin page or another appropriate page
    return redirect(url_for('admin'))

@app.route('/drafts')
def drafts():
    return render_template('drafts.html')

@app.route('/get-names', methods=['GET'])
def get_names():
    df = pd.read_csv('pending_submissions.csv')
    names = df['Name'].tolist()
    return jsonify(names)

@app.route('/get-record', methods=['GET'])
def get_record():
    name = request.args.get('name')
    df = pd.read_csv('pending_submissions.csv')
    record = df[df['Name'] == name].to_dict(orient='records')
    return jsonify(record[0] if record else {})

@app.route('/update-record', methods=['POST'])
def update_record():
    data = request.json
    df = pd.read_csv('pending_submissions.csv')

    # Extract the name and the new notes value from the request data
    name = data.get('Name')
    new_notes = data.get('Notes')

    # Check if the name exists in the DataFrame
    if name in df['Name'].values:
        # Update the 'Notes' column for the row where 'Name' matches
        df.loc[df['Name'] == name, 'Notes'] = new_notes
        df.to_csv('pending_submissions.csv', index=False)
        return jsonify({"status": "success"})
    else:
        return jsonify({"error": "Name not found"}), 404



@app.route('/logout')
def logout():
    session.pop('username', None)  # Remove the username from the session
    return redirect(url_for('login'))

BASE_DIR = 'Demerits'  # Base directory for the files

def allowed_file(directory):
    return os.path.commonprefix([os.path.abspath(directory), os.path.abspath(BASE_DIR)]) == os.path.abspath(BASE_DIR)

@app.route('/directory')
def directory():
    return render_template('directory.html')

@app.route('/grades/<grade>')
def get_students(grade):
    directory = os.path.join(BASE_DIR, f'Grade_{grade}')
    if not allowed_file(directory):
        return jsonify({"error": "Access denied"}), 403
    
    try:
        students = [name for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]
        return jsonify(students)
    except FileNotFoundError:
        return jsonify([])

@app.route('/files', methods=['GET'])
def list_files():
    grade = request.args.get('grade')
    name = request.args.get('name')
    date = request.args.get('date')

    target_dir = os.path.join(BASE_DIR, f'Grade_{grade}', name) if grade and name else BASE_DIR
    print(f"Target directory: {target_dir}")  # Debug print

    if not allowed_file(target_dir):
        return jsonify({"error": "Access denied"}), 403

    files = []
    for root, dirs, filenames in os.walk(target_dir):
        for filename in filenames:
            if date and date not in filename:
                continue
            full_path = os.path.relpath(os.path.join(root, filename), BASE_DIR)
            print(f"Adding file: {full_path}")  # Debug print
            files.append(full_path)

    return jsonify(files)

@app.route('/download', methods=['POST'])
def download_files():
    try:
        data = request.get_json()
        files = data['files']
        print("Received files:", files)  # Debug output

        if len(files) == 1:
            file_path = os.path.join(BASE_DIR, files[0])
            directory = os.path.dirname(file_path)
            filename = os.path.basename(file_path)
            print("Attempting to send file from path:", file_path)  # More debug output

            if os.path.exists(file_path) and os.path.isfile(file_path):
                return send_from_directory(directory, filename, as_attachment=True)
            else:
                print("File does not exist:", file_path)  # File existence debug output
                return jsonify({"error": "File does not exist"}), 404

    except Exception as e:
        traceback.print_exc()  # Print stack trace for detailed debug info
        return jsonify({"error": "Internal Server Error", "message": str(e)}), 500
    
    
    # For multiple files, create a zip
    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file in files:
            full_path = os.path.join(BASE_DIR, file)
            if os.path.exists(full_path) and os.path.isfile(full_path):
                zip_file.write(full_path, arcname=os.path.basename(full_path))
            else:
                return jsonify({"error": f"File not found: {file}"}), 404
    zip_stream.seek(0)

    return Response(zip_stream.getvalue(),
                    mimetype='application/zip',
                    headers={"Content-Disposition": "attachment;filename=download.zip"})


def get_tables(mdb_path):
    conn_str = (
        r'DRIVER={MDBTools};'
        f'DBQ={mdb_path};'
    )
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    cursor.tables()
    tables = [row.table_name for row in cursor if row.table_type == 'TABLE']
    conn.close()
    return tables

def read_mdb_table(mdb_path, table_name):
    conn_str = (
        r'DRIVER={MDBTools};'
        f'DBQ={mdb_path};'
    )
    conn = pyodbc.connect(conn_str)
    query = f'SELECT * FROM {table_name}'
    df = pd.read_sql(query, conn)
    conn.close()
    return df

@app.route('/db', methods=['GET', 'POST'])
def db():
    mdb_path = '/home/mq/kdb.mdb'  # Update with your MDB file path
    tables = get_tables(mdb_path)
    
    if request.method == 'POST':
        table_name = request.form['table_name']
        try:
            df = read_mdb_table(mdb_path, table_name)
            csv_path = f'/path/to/save/{table_name}.csv'  # Update with your desired CSV file path
            df.to_csv(csv_path, index=False)
            message = f"Table {table_name} exported to CSV successfully!"
        except Exception as e:
            message = f"Error: {str(e)}"
        return render_template('db.html', tables=tables, message=message)

    return render_template('db.html', tables=tables)

if __name__ == '__main__':
    app.run(debug=True, host='100.105.121.42', port=8081)
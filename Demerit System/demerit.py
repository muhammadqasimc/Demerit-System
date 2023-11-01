import base64
import io
import uuid
from zipfile import ZipFile
from flask import Flask, jsonify, render_template, request, redirect, url_for, session, send_from_directory
import pandas as pd
import os
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from playsound import playsound
import csv  # Added for CSV file operations
from flask import send_file, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas  # Add this import statement
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from fpdf import FPDF




app = Flask(__name__)

# Constants for file paths
STUDENT_DATA_CSV = 'student_data.csv'
PENDING_SUBMISSIONS_CSV = 'pending_submissions.csv'

# Lists to store submissions
submissions_pending_approval = []
approved_submissions = []

# Declare the df variable as global
df = None

def load_data():
    global df
    df = pd.read_csv(STUDENT_DATA_CSV)

# Call the load_data function to load the CSV data when the app starts
load_data()

# Function to load pending submissions from a CSV file
def load_pending_submissions():
    pending_submissions = []
    try:
        with open(PENDING_SUBMISSIONS_CSV, mode='r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                pending_submissions.append(row)
    except FileNotFoundError:
        # If the file doesn't exist, start with an empty list
        pass
    return pending_submissions

# Function to save pending submissions to a CSV file
def save_pending_submissions(submissions):
    with open(PENDING_SUBMISSIONS_CSV, mode='w', newline='') as file:
        fieldnames = ["Name", "Grade", "Date", "Notes", 'Username', "Offenses", "StudentSignature", "TeacherSignature", "WitnessSignature"]
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for submission in submissions:
            writer.writerow(submission)

# Load pending submissions when the app starts
submissions_pending_approval = load_pending_submissions()

# Dummy user data for demonstration purposes
users = {
    'JAMESON A -4': {
        'password': 'password123',
        'is_admin': True
    },
    'KGONGWANA E -20': {
        'password': 'userpassword',
        'is_admin': False
    },
    'VARGHESE S -16': {
        'password': 'randompassword1',
        'is_admin': False
    },
    'JOHN B -14': {
        'password': 'randompassword2',
        'is_admin': False
    },
    'JOUBERT R -2': {
        'password': 'randompassword3',
        'is_admin': False
    },
    'JAMESON C -12': {
        'password': 'randompassword4',
        'is_admin': False
    },
    'BESTER N -17': {
        'password': 'randompassword5',
        'is_admin': False
    },
    'CLAASSENS CH -22': {
        'password': 'randompassword6',
        'is_admin': False
    },
    'VAN EEDEN B -38': {
        'password': 'randompassword7',
        'is_admin': False
    },
    'GELDENHUYS SM -44': {
        'password': 'randompassword8',
        'is_admin': False
    },
    'MOSENTHAL V D -46': {
        'password': 'randompassword9',
        'is_admin': False
    },
    'BEZUIDENHOUT M -47': {
        'password': 'randompassword10',
        'is_admin': False
    },
    'STRYDOM F -48': {
        'password': 'randompassword11',
        'is_admin': False
    },
    'BRUWER B -51': {
        'password': 'randompassword12',
        'is_admin': False
    },
    'FOURIE L -49': {
        'password': 'randompassword13',
        'is_admin': False
    },
    'JANKOWITZ J -52': {
        'password': 'randompassword14',
        'is_admin': False
    },
    'TAYOB N -53': {
        'password': 'randompassword15',
        'is_admin': False
    },
    'PATTASSERIL BABU B -55': {
        'password': 'randompassword16',
        'is_admin': False
    },
    'GELDENHUYS H D -56': {
        'password': 'randompassword17',
        'is_admin': False
    },
    'GOOSEN C -57': {
        'password': 'randompassword18',
        'is_admin': False
    },
    'REDIKER L -58': {
        'password': 'randompassword19',
        'is_admin': False
    },
    'JONKER F -59': {
        'password': 'randompassword20',
        'is_admin': False
    },
    'DU TOIT WD -60': {
        'password': 'randompassword21',
        'is_admin': False
    },
    'admin': {
        'password': '123',
        'is_admin': True
    }
}

# Set a secret key for session management
app.secret_key = os.urandom(24)

def is_logged_in():
    return 'username' in session


@app.route('/')
def index():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()
        
        return render_template('index.html', grades=grades, namesByGrade=names_by_grade)
    else:
        # Redirect to the login page if the user is not logged in
        return redirect(url_for('login'))

@app.route('/green')
def green():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()
        return render_template('green.html', grades=grades, namesByGrade=names_by_grade)
    else:
        return redirect(url_for('login'))

@app.route('/yellow')
def yellow():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()
        return render_template('yellow.html', grades=grades, namesByGrade=names_by_grade)
    else:
        return redirect(url_for('login'))
    

@app.route('/pink')
def pink():
    if 'username' in session:
        df['Grade'] = df['Grade'].astype(int)
        
        # Get unique grade values
        grades = sorted(df['Grade'].unique())
        
        # Create a dictionary to store names grouped by grade
        names_by_grade = {}
        for grade in grades:
            names_by_grade[grade] = df[df['Grade'] == grade]['Name'].tolist()
        return render_template('pink.html', grades=grades, namesByGrade=names_by_grade)
    else:
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username in users and password == users[username]['password']:
            session['username'] = username
            return redirect(url_for('index'))

    return render_template('login.html')

@app.route('/select_name', methods=['POST'])
def select_name():
    selected_name = request.form['name']
    
    # Filter data by the selected name and non-empty Date and Offenses
    filtered_data = df[(df['Name'] == selected_name) & ~df['Date'].isnull() & ~df['Offenses'].isnull()]
    
    return render_template('select_name.html', data=filtered_data)

@app.route('/get_names_by_grade/<int:selected_grade>')
def get_names_by_grade(selected_grade):
    names = df[df['Grade'] == selected_grade]['Name'].tolist()
    return jsonify(names)

# Define the directory where signature images will be saved
signature_folder = 'signatures'

@app.route('/submit_form', methods=['POST'])
def submit_form():
    selected_grade = request.form['grade']
    selected_name = request.form['name']
    selected_offenses = request.form.getlist('offense')
    current_date = datetime.now().strftime('%Y-%m-%d')
    notes = request.form['notes']  # Retrieve the notes from the form
    username = session['username']

     # Handle the student's signature
    student_signature_data_url = request.form['student_signature']
    if student_signature_data_url:
        # Generate a unique filename for the student's signature image
        student_signature_filename = f'{selected_name}_{current_date}.png'
        student_signature_path = os.path.join(signature_folder, student_signature_filename)

        # Convert the data URL back to an image and save it as a PNG
        student_signature_image_data = student_signature_data_url.split(',')[1]
        student_signature_image_binary = base64.b64decode(student_signature_image_data)
        with open(student_signature_path, 'wb') as student_signature_file:
            student_signature_file.write(student_signature_image_binary)

    # Handle the teacher's signature (similar code as above)
    teacher_signature_data_url = request.form['teacher_signature']
    if teacher_signature_data_url:
        teacher_signature_filename = f'{username}_{current_date}.png'
        teacher_signature_path = os.path.join(signature_folder, teacher_signature_filename)
        teacher_signature_image_data = teacher_signature_data_url.split(',')[1]
        teacher_signature_image_binary = base64.b64decode(teacher_signature_image_data)
        with open(teacher_signature_path, 'wb') as teacher_signature_file:
            teacher_signature_file.write(teacher_signature_image_binary)


    # Handle the witness's signature (similar code as above)
    witness_signature_data_url = request.form['witness_signature']
    if witness_signature_data_url:
        witness_signature_filename = f'witness_signature_{datetime.now().strftime("%Y%m%d%H%M%S")}.png'
        witness_signature_path = os.path.join(signature_folder, witness_signature_filename)
        witness_signature_image_data = witness_signature_data_url.split(',')[1]
        witness_signature_image_binary = base64.b64decode(witness_signature_image_data)
        with open(witness_signature_path, 'wb') as witness_signature_file:
            witness_signature_file.write(witness_signature_image_binary)


    # Create a new submission dictionary
    new_submission = {
        'Name': selected_name,
        'Grade': selected_grade,
        'Date': current_date,
        'Offenses': ', '.join(selected_offenses),
        'Notes': notes,
        'Username' : username
    }

    # Append the submission to the pending approval list
    submissions_pending_approval.append(new_submission)

    # Save pending submissions to the CSV file
    save_pending_submissions(submissions_pending_approval)

    # Notify with a sound
    playsound('static/ding.mp3')

    return redirect(url_for('index'))

@app.route('/admin')
def admin():
    if 'username' in session:
        username = session['username']
        if users[username]['is_admin']:
            return render_template('admin.html', submissions=submissions_pending_approval)
        else:
            return 'You are not authorized to access this page.'
    return redirect(url_for('login'))

@app.route('/approve_submission/<int:index>')
def approve_submission(index):
    if 0 <= index < len(submissions_pending_approval):
        submission = submissions_pending_approval.pop(index)
        approved_submissions.append(submission)

        # Create directory structure for the grade
        grade_folder = os.path.join('Demerits', f'Grade_{submission["Grade"]}')
        os.makedirs(grade_folder, exist_ok=True)

        # Get the current time in HH:MM format
        current_time = datetime.now().strftime("%H:%M")

        logo_image = Image.open('static/Logo.png')
        logo_image = logo_image.convert('RGB')  # Convert to RGB format (optional)
        logo_image.save('static/Logo_non_interlaced1.png', 'PNG', optimize=True)

        logo_filename = 'static/Logo_non_interlaced1.png'
        logo_width = 130  # Width of the logo in millimeters
        logo_height = 40  # Height of the logo in millimeters


        # Use the current time in the filename
        pdf_filename = f'{submission["Name"]}_{submission["Date"]}_{current_time}.pdf'
        pdf_path = os.path.join(grade_folder, pdf_filename)

        # Create a PDF with A4 size (210x297 mm) in landscape mode
        pdf = FPDF(format='A4', unit='mm')
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)  # Automatic page break with a 15mm margin
        pdf.set_font("Arial", size=12)
        
        # Calculate the x-coordinate to center the logo horizontally
        logo_x = (pdf.w - logo_width) / 2

        # Load and embed the logo image
        pdf.image(logo_filename, x=logo_x, y=pdf.get_y(), w=logo_width, h=logo_height)  # Adjust position and size

        # Move down to create space for the logo
        pdf.ln(logo_height)


        # Adjusted positions and sizes for A4 portrait
        pdf.cell(0, 10, txt="Demerits Form", ln=True, align='C')  # Background fill
        pdf.ln(10)  # Add some space

        # Left Column
        pdf.set_fill_color(200, 200, 200)  # Background fill color
        column_width = 90  # Adjust the column width as needed
        column_width1 = 25  # Adjust the column width as needed

        pdf.cell(column_width1, 10, txt="Name:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Name'], ln=True)

        pdf.cell(column_width1, 10, txt="Grade:", fill=True)
        pdf.cell(column_width, 10, txt=str(submission['Grade']), ln=True)

        pdf.cell(column_width1, 10, txt="Date:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Date'], ln=True)

        pdf.cell(column_width1, 10, txt="Notes:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Notes'], ln=True)

        pdf.cell(column_width1, 10, txt="Educator:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Username'], ln=True)

        pdf.cell(column_width1, 10, txt="Offenses:", fill=True)
        pdf.cell(column_width, 10, txt=submission['Offenses'], ln=True)  # Multiline cell for offenses

        pdf.ln(10)  # Add some space

        # Title for the student's signature cell
        pdf.set_font("Arial", style='B', size=12)
        pdf.cell(0, 10, txt="Student Signature", ln=True, align='C')

        # Create a cell for the student's signature
        pdf.set_fill_color(255, 255, 255)  # White background
        pdf.cell(0, 30, txt="", border=1, ln=True)  # Cell for the signature with borders

        # Load and embed the student's signature image within the cell
        student_signature_filename = f'{submission["Name"]}_{submission["Date"]}.png'
        student_signature_path = os.path.join(signature_folder, student_signature_filename)
        pdf.image(student_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() - 25, w=0, h=20)  # Adjust position and size

        # Title for the teacher's signature cell
        pdf.set_font("Arial", style='B', size=12)
        pdf.cell(0, 10, txt="Teacher Signature", ln=True, align='C')

        # Create a cell for the teacher's signature
        pdf.set_fill_color(255, 255, 255)  # White background
        pdf.cell(0, 30, txt="", border=1, ln=True)  # Cell for the signature with borders

        # Load and embed the teacher's signature image within the cell
        teacher_signature_filename = f'{submission["Username"]}_{submission["Date"]}.png'
        teacher_signature_path = os.path.join(signature_folder, teacher_signature_filename)
        pdf.image(teacher_signature_path, x=pdf.get_x() + 5, y=pdf.get_y() - 25, w=0, h=20)  # Adjust position and size


        # # Title for the witness's signature
        # pdf.set_font("Arial", style='B', size=12)
        # pdf.cell(20190, 10, txt="Witness Signature", ln=True, align='C')

        # # Create a box for the witness's signature
        # pdf.set_fill_color(255, 255, 255)  # White background
        # pdf.rect(10, pdf.get_y(), 190, 40, style='F')  # Rectangle for the signature

        # # Load and embed the witness's signature image in the box
        # witness_signature_filename = f'witness_signature_{submission["Date"]}.png'
        # witness_signature_path = os.path.join(signature_folder, witness_signature_filename)
        # pdf.image(witness_signature_path, x=15, y=pdf.get_y() + 5, w=180, h=30)  # Adjust position and size

        # Save the PDF to the file
        pdf.output(pdf_path)

        # Save pending submissions to the CSV file
        save_pending_submissions(submissions_pending_approval)

    return redirect(url_for('admin'))

@app.route('/reject_submission/<int:index>')
def reject_submission(index):
    if 0 <= index < len(submissions_pending_approval):
        submission = submissions_pending_approval.pop(index)
        # Perform any rejection action here (e.g., delete the submission)
        # Save pending submissions to the CSV file
        save_pending_submissions(submissions_pending_approval)
        # Notify with a sound

    return redirect(url_for('admin'))

@app.route('/logout')
def logout():
    session.pop('username', None)  # Remove the username from the session
    return redirect(url_for('login'))
    

@app.route('/directory')
def directory():
    # Specify the directory where the PDF files are located
    pdf_directory = 'Demerits'

    # Create a list to store the available PDF files
    pdf_file_list = []

    # Loop through the grade directories and list PDF files
    for grade_folder in os.listdir(pdf_directory):
        if os.path.isdir(os.path.join(pdf_directory, grade_folder)):
            grade_path = os.path.join(pdf_directory, grade_folder)
            for pdf_file in os.listdir(grade_path):
                if pdf_file.endswith('.pdf'):
                    # Construct the full file path
                    file_path = os.path.join(grade_path, pdf_file)
                    # Create a dictionary with file information
                    pdf_info = {
                        'grade_folder': grade_folder,
                        'file_name': pdf_file,
                        'file_path': file_path,
                    }
                    pdf_file_list.append(pdf_info)

    # Get unique grade values for the grade filter dropdown
    grades = sorted(set(info['grade_folder'] for info in pdf_file_list))

    # Get unique student names for the student filter dropdown
    students = sorted(set(info['file_name'].split('_')[0] for info in pdf_file_list))

    return render_template('directory.html', pdf_files=pdf_file_list, grades=grades, students=students,
                           filtered_pdf_files=pdf_file_list)

@app.route('/filter_directory', methods=['POST'])
def filter_directory():
    selected_grade = request.form.get('selected_grade')
    selected_student = request.form.get('selected_student')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    # Specify the directory where the PDF files are located
    pdf_directory = 'Demerits'

    # Create a list to store the available PDF files
    pdf_file_list = []

    # Loop through the grade directories and list PDF files
    for grade_folder in os.listdir(pdf_directory):
        if os.path.isdir(os.path.join(pdf_directory, grade_folder)):
            grade_path = os.path.join(pdf_directory, grade_folder)
            for pdf_file in os.listdir(grade_path):
                if pdf_file.endswith('.pdf'):
                    # Construct the full file path
                    file_path = os.path.join(grade_path, pdf_file)
                    # Check if the file matches the filter criteria
                    file_info = {
                        'grade_folder': grade_folder,
                        'file_name': pdf_file,
                        'file_path': file_path,
                    }
                    if (not selected_grade or file_info['grade_folder'] == selected_grade) and \
                       (not selected_student or file_info['file_name'].startswith(selected_student)) and \
                       (not start_date or file_info['file_name'].split('_')[1] >= start_date) and \
                       (not end_date or file_info['file_name'].split('_')[1] <= end_date):
                        pdf_file_list.append(file_info)

    # Get unique grade values for the grade filter dropdown
    grades = sorted(set(info['grade_folder'] for info in pdf_file_list))

    # Get unique student names for the student filter dropdown
    students = sorted(set(info['file_name'].split('_')[0] for info in pdf_file_list))

    return render_template('directory.html', pdf_files=pdf_file_list, grades=grades, students=students,
                           filtered_pdf_files=pdf_file_list)

@app.route('/bulk_download_directory', methods=['POST'])
def bulk_download_directory():
    selected_files = request.form.getlist('selected_files')

    # Check if any files are selected for download
    if not selected_files:
        return "No files selected for download."

    # Create a zip file containing the selected PDF files
    zip_filename = 'bulk_download.zip'
    with ZipFile(zip_filename, 'w') as zipf:
        for file_path in selected_files:
            # Make sure the selected file exists and is allowed for download
            if os.path.exists(file_path):
                zipf.write(file_path, os.path.basename(file_path))

    # Send the zip file for download
    return send_file(zip_filename, as_attachment=True)


@app.route('/pdf_files', methods=['GET', 'POST'])
def pdf_files():
    # Create a list to store the available PDF files
    pdf_file_list = []

    # Specify the directory where the PDF files are located (same as CSV)
    pdf_directory = 'Demerits'

    # Loop through the grade directories and list PDF files
    for grade_folder in os.listdir(pdf_directory):
        if os.path.isdir(os.path.join(pdf_directory, grade_folder)):
            grade_path = os.path.join(pdf_directory, grade_folder)
            for pdf_file in os.listdir(grade_path):
                if pdf_file.endswith('.pdf'):
                    # Construct the full file path
                    file_path = os.path.join(grade_path, pdf_file)
                    # Create a dictionary with file information
                    pdf_info = {
                        'grade_folder': grade_folder,
                        'file_name': pdf_file,
                        'file_path': file_path,
                    }
                    pdf_file_list.append(pdf_info)

    # Get unique grade values for the grade filter dropdown
    grades = sorted(set(info['grade_folder'] for info in pdf_file_list))

    # Get unique student names for the student filter dropdown
    students = sorted(set(info['file_name'].split('_')[0] for info in pdf_file_list))

    # Initialize the filtered file list
    filtered_pdf_files = pdf_file_list

    if request.method == 'POST':
        selected_grade = request.form.get('selected_grade')
        selected_student = request.form.get('selected_student')

        # Apply grade filter
        if selected_grade:
            filtered_pdf_files = [info for info in filtered_pdf_files if info['grade_folder'] == selected_grade]

        # Apply student filter
        if selected_student:
            filtered_pdf_files = [info for info in filtered_pdf_files if info['file_name'].startswith(selected_student)]

    return render_template('pdf_files.html', pdf_files=pdf_file_list, grades=grades, students=students,
                           filtered_pdf_files=filtered_pdf_files)

@app.route('/bulk_download_pdf', methods=['POST'])
def bulk_download_pdf():
    selected_files = request.form.getlist('selected_files')

    # Check if any files are selected for download
    if not selected_files:
        return "No files selected for download."

    # Specify the directory where the PDF files are located
    pdf_directory = 'Demerits'

    # Create a zip file containing the selected PDF files
    zip_filename = 'bulk_download_pdf.zip'
    with ZipFile(zip_filename, 'w') as zipf:
        for file_name in selected_files:
            # Construct the full file path
            file_path = os.path.join(pdf_directory, file_name)
            
            # Make sure the selected file exists and is allowed for download
            if os.path.exists(file_path):
                zipf.write(file_path, file_name)  # Use file_name as the archive name

    # Send the zip file for download
    return send_file(zip_filename, as_attachment=True)


if __name__ == '__main__':
    app.run(host='172.20.10.3', port=8080)